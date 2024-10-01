import openpyxl as xl
import random



book = xl.load_workbook("./practice/toeic_word.xlsx")
sheet = book.active
oCount = 0
newOCount = 0
wordDictionary = {}
newWordDictionary = {}



# 함수로 처리하면 더 깔끔할거 같아서 함수로 묶음
def createNewFile(finalWordList) :
    new_book = xl.Workbook()
    new_sheet = new_book.active
    new_sheet.column_dimensions['A'].width = 15
    new_sheet.column_dimensions['B'].width = 80

    for row, rowVal in enumerate(finalWordList) :
        cE = new_sheet.cell(row+1, 1)
        cK = new_sheet.cell(row+1, 2)
        cE.value = rowVal[0]
        cK.value = rowVal[1]

    new_book.save("./practice/new_word_list.xlsx")



try :
    newBook = xl.load_workbook("./practice/new_word_list.xlsx")
    newSheet = newBook.active
    for r in range(1, 1223) : # 값이 들어가 있는 셀의 최대값이 어디인지 찾아야함. 일단 임의의 값으로 1223을 줌
        newCheck = newSheet.cell(row = r, column = 3).value
        if newCheck == "ㅇ" :
            newOCount += 1
        else :
            newWordDictionary[newSheet.cell(row = r, column = 1).value] = newSheet.cell(row = r, column = 2).value
            
    newWordList = list(newWordDictionary.items())
     
    random.shuffle(newWordList)
    
    percentage = round((1222-newOCount)/1222*100, 2) # 여기 고쳐야 함, 전체 갯수를 몰라서 1222 넣었는데 이렇게 하면 오류 날거임

    # print(list(enumerate(wordList)))
    print("{0}/1222 {1}%".format((1222-newOCount), percentage)) # 46번째 줄과 마찬가지로 여기도 고쳐야함

    createNewFile(newWordList)
    
    
    
except FileNotFoundError:
    for r in range(2, 1223) :
        check = sheet.cell(row = r, column = 4).value
        if check == "ㅇ" : 
            oCount += 1
        else :
            wordDictionary[sheet.cell(row = r, column = 2).value] = sheet.cell(row = r, column = 3).value
    
    wordList = list(wordDictionary.items()) # items()함수를 써야지 키와 값이 함께 튜플에 담김
    wordList = wordList

    random.shuffle(wordList)
    
    percentage = round(oCount/1222*100, 2)
    
    # print(list(enumerate(wordList)))
    print("{0}/1222 {1}%".format(oCount, percentage))
    
    createNewFile(wordList)
    
# 계속 갯수가 한 개 차이 나는데 이건 아직 원인을 모르겠음. 수정 필요
